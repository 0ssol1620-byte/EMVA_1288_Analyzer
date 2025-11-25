# -*- coding: utf-8 -*-
"""
EMVA 1288 Analyzer — Compact Professional UI, Robust Windows Taskbar Icon

예)
  python main.py --theme auto
  python main.py --theme dark --icon assets/vieworks.ico
"""
import sys
import os
import time
import argparse
import contextlib
import numpy as np
import openpyxl
from contextlib import suppress
from typing import Optional, List, Dict, Tuple
from pathlib import Path
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QTableWidget, QTableWidgetItem,
    QProgressBar, QGraphicsView, QGraphicsScene, QSpinBox, QMessageBox,
    QFileDialog, QGroupBox, QFormLayout, QComboBox, QHeaderView, QStyle,
    QSplitter, QStatusBar, QFrame, QSizePolicy, QScrollArea, QProxyStyle
)
from PyQt5.QtGui import QImage, QPixmap, QIcon, QFont, QPalette, QColor, QPainter
from PyQt5.QtCore import Qt, QTimer, pyqtSlot, QEventLoop, QSize

# 실제 프로젝트 경로에서 import
from core.camera_facade import CxpCamera
from core.lightbox_controller import LightBoxController
from workers.measurement_worker import MeasurementWorker, SaturationWorker, VALID_BAYER
from workers.dark_data_worker import DarkDataWorker


# -------------------------------
# CLI
# -------------------------------
def parse_args(argv=None):
    p = argparse.ArgumentParser(description="EMVA 1288 Analyzer")
    p.add_argument("--theme", default="dark", choices=["auto", "light", "dark"], help="UI theme")
    p.add_argument("--icon", default="", help="Path to .ico/.png for window/taskbar icon")
    p.add_argument("--no-header", action="store_true", help="Hide custom header bar")
    return p.parse_args(argv)


def _template_candidates(filename: str = "Template_format.xlsx") -> List[Path]:
    """
    템플릿을 찾을 수 있는 모든 후보 경로를 '순서대로' 반환.
    1) PyInstaller 런타임 임시폴더(_MEIPASS)         ← 번들에 포함했을 때
    2) 실행파일(exe) 폴더                            ← 배포 exe 옆에 둔 경우
    3) 스크립트(__file__) 폴더                        ← 개발/디버그 실행
    4) 현재 작업 디렉터리(CWD)                        ← 기타 경우(IDE 설정 등)
    """
    bases: List[Path] = []
    # 1) _MEIPASS
    if hasattr(sys, "_MEIPASS") and sys._MEIPASS:
        with contextlib.suppress(Exception):
            bases.append(Path(sys._MEIPASS).resolve())
    # 2) exe 폴더
    with contextlib.suppress(Exception):
        bases.append(Path(sys.executable).resolve().parent)
    # 3) 스크립트 폴더
    with contextlib.suppress(Exception):
        bases.append(Path(__file__).resolve().parent)
    # 4) CWD
    bases.append(Path.cwd())

    # 중복 제거(앞쪽 우선)
    uniq: List[Path] = []
    seen = set()
    for b in bases:
        if not b:
            continue
        nb = b.resolve()
        if nb in seen:
            continue
        seen.add(nb)
        uniq.append(nb)
    return [p / filename for p in uniq]

def _template_path(filename: str = "Template_format.xlsx") -> Path:
    """
    후보 경로를 순서대로 검사해서 '실제로 존재하는' 첫 경로를 반환.
    모두 없으면 2순위(exe 폴더) 경로를 기본 반환(에러 메시지에 후보 리스트를 출력).
    """
    cands = _template_candidates(filename)
    for p in cands:
        if p.exists():
            return p
    # 아무것도 없으면 exe 폴더를 디폴트로 반환
    # (index 1이 항상 exe 폴더가 되도록 _template_candidates에서 순서를 잡아 둠)
    return cands[1] if len(cands) > 1 else cands[0]
# -------------------------------
# Windows: AppUserModelID & Icon
# -------------------------------
AUMID = "com.vieworks.DefectPixelFinder"  # 임의 가능. 단, 항상 동일 값 사용 권장.

def _set_windows_app_user_model_id(app_id: str) -> None:
    """Windows 작업표시줄에 아이콘 정확 반영 (Qt 생성 전 호출 필수)."""
    if sys.platform.startswith("win"):
        try:
            import ctypes
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(app_id)
            print(f"[AUMID] {app_id}")
        except Exception as e:
            print(f"[AUMID] failed: {e}")

def _force_window_icon_win(hwnd, ico_path: str) -> None:
    """런타임 + 클래스 아이콘 모두 설정 (작업표시줄/Alt-Tab 완전 일치)."""
    if not (sys.platform.startswith("win") and os.path.exists(ico_path)):
        return
    import ctypes
    user32 = ctypes.windll.user32

    LR_LOADFROMFILE = 0x0010
    LR_DEFAULTSIZE  = 0x0040
    IMAGE_ICON      = 1
    WM_SETICON      = 0x0080
    ICON_SMALL, ICON_BIG = 0, 1

    # 클래스 아이콘 세팅용 상수
    GCLP_HICON      = -14
    GCLP_HICONSM    = -34
    SetClassLongPtr = getattr(user32, "SetClassLongPtrW", None) or user32.SetClassLongW

    # 아이콘 로드(기본 사이즈 자동 선택)
    hicon = user32.LoadImageW(None, ico_path, IMAGE_ICON, 0, 0, LR_LOADFROMFILE | LR_DEFAULTSIZE)
    if not hicon:
        return

    # 런타임 아이콘
    user32.SendMessageW(hwnd, WM_SETICON, ICON_BIG,   hicon)
    user32.SendMessageW(hwnd, WM_SETICON, ICON_SMALL, hicon)

    # 클래스 아이콘(Alt-Tab/작업표시줄 일관성)
    try:
        SetClassLongPtr(hwnd, GCLP_HICON,   hicon)
        SetClassLongPtr(hwnd, GCLP_HICONSM, hicon)
    except Exception:
        pass

def resource_path(*paths) -> str:
    """
    리소스(템플릿/아이콘)를 '실행 컨텍스트' 기준의 안정적인 절대경로로 변환.
    우선순위:
      1) PyInstaller 런타임 임시폴더(sys._MEIPASS)
      2) 동작중인 실행파일/런처(sys.executable, sys.argv[0])의 폴더
      3) 이 파일(__file__)의 폴더
      4) 최후의 보루: 현재 작업 디렉터리(CWD)
    """
    # 1) PyInstaller 런타임 임시폴더
    if hasattr(sys, "_MEIPASS") and sys._MEIPASS:
        base = sys._MEIPASS
    else:
        # 2) 실행파일/런처 기준
        try:
            base = os.path.dirname(os.path.abspath(sys.executable))
            if not base or not os.path.exists(base):
                raise RuntimeError
        except Exception:
            # 3) 소스 파일 기준
            try:
                base = os.path.dirname(os.path.abspath(__file__))
            except Exception:
                # 4) CWD
                base = os.path.abspath(".")

    return os.path.normpath(os.path.join(base, *paths))

def _normalize_path(p: str) -> str:
    return os.path.normpath(os.path.abspath(p))

def load_app_icon(icon_hint: str = "") -> Tuple[Optional[QIcon], Optional[str]]:
    """
    아이콘 로더: 우선순위
    1) --icon 인자(존재하면)
    2) 환경변수 VIEWORKS_ICON
    3) 실행 스크립트/패키지 루트: assets/vieworks.ico, vieworks.ico (+레거시 app_icon.*)
    """
    candidates: List[str] = []
    if icon_hint:
        candidates.append(icon_hint)

    env_icon = os.environ.get("VIEWORKS_ICON", "")
    if env_icon:
        candidates.append(env_icon)

    script_dir = os.path.dirname(os.path.abspath(sys.argv[0] if sys.argv and sys.argv[0] else __file__))
    # vieworks.ico 우선
    for rel in [
        os.path.join(script_dir, "assets", "vieworks.ico"),
        os.path.join(script_dir, "vieworks.ico"),
        resource_path("assets", "vieworks.ico"),
        resource_path("vieworks.ico"),
        # 하위 호환: 기존 이름도 탐색
        os.path.join(script_dir, "assets", "app_icon.ico"),
        os.path.join(script_dir, "app_icon.ico"),
        os.path.join(script_dir, "app_icon.png"),
        resource_path("assets", "app_icon.ico"),
        resource_path("app_icon.ico"),
        resource_path("app_icon.png"),
    ]:
        candidates.append(rel)

    seen = set()
    for c in candidates:
        if not c:
            continue
        cp = _normalize_path(c)
        if cp in seen:
            continue
        seen.add(cp)
        if os.path.exists(cp):
            ic = QIcon(cp)
            if not ic.isNull():
                print(f"[ICON] loaded: {cp}")
                return ic, cp
            else:
                print(f"[ICON] found but invalid: {cp}")

    print("[ICON] not found; using default")
    return None, None

def detect_windows_dark() -> bool:
    """Windows 다크 모드 감지 (AppsUseLightTheme == 0)."""
    if not sys.platform.startswith("win"):
        return False
    try:
        import winreg
        with winreg.OpenKey(
            winreg.HKEY_CURRENT_USER,
            r"Software\Microsoft\Windows\CurrentVersion\Themes\Personalize"
        ) as k:
            v, _ = winreg.QueryValueEx(k, "AppsUseLightTheme")
            return int(v) == 0
    except Exception:
        return False


# -------------------------------
# Constants & Compact metrics
# -------------------------------
DEFAULT_LB_MAX_UNITS = 65535
SHEET_LINEARITY = "Linearity Data"
SHEET_DARK      = "Linearity Data(Exposure)"

CONTROL_H = 28
RADIUS    = 8


# -------------------------------
# Helpers (math)
# -------------------------------
def compute_light_steps(saturation_level: int, num_steps: int) -> List[int]:
    sat = int(max(0, saturation_level))
    n = int(max(1, num_steps))
    if n == 1:
        return [sat] if sat > 0 else [0]
    xs = np.linspace(0, sat, n, endpoint=True, dtype=np.float64)
    steps = np.rint(xs).astype(np.int64)
    for i in range(1, n):
        if steps[i] <= steps[i - 1]:
            steps[i] = steps[i - 1] + 1
    over = steps[-1] - sat
    if over > 0:
        for i in range(n - 1, -1, -1):
            if over == 0:
                break
            target = max(0 if i == 0 else steps[i - 1] + 1, steps[i] - 1)
            dec = steps[i] - target
            if dec > 0:
                take = min(dec, over)
                steps[i] -= take
                over -= take
    steps[0] = 0
    steps[-1] = sat
    return steps.astype(int).tolist()


# -------------------------------
# G-only preview utils
# -------------------------------
def bayer_green_compact(raw_bayer: np.ndarray, pattern: str) -> Optional[np.ndarray]:
    if raw_bayer is None or raw_bayer.ndim != 2 or pattern not in VALID_BAYER:
        return None
    H, W = raw_bayer.shape
    H2, W2 = (H // 2) * 2, (W // 2) * 2
    r = raw_bayer[:H2, :W2]
    p = pattern.upper()
    if p == "RGGB":
        Gr = r[0::2, 1::2]; Gb = r[1::2, 0::2]
    elif p == "BGGR":
        Gr = r[1::2, 0::2]; Gb = r[0::2, 1::2]
    elif p == "GRBG":
        Gr = r[0::2, 0::2]; Gb = r[1::2, 1::2]
    elif p == "GBRG":
        Gr = r[1::2, 1::2]; Gb = r[0::2, 0::2]
    else:
        return None
    return ((Gr.astype(np.float64) + Gb.astype(np.float64)) * 0.5).astype(r.dtype)

def upsample_nn2x(img: np.ndarray) -> np.ndarray:
    if img is None or img.ndim != 2:
        return img
    return np.repeat(np.repeat(img, 2, axis=0), 2, axis=1)

def to_g_mono(frame: np.ndarray, pixel_format: str, pattern: Optional[str]) -> np.ndarray:
    if frame is None:
        return None
    pf = (pixel_format or "").upper()
    if frame.ndim == 2 and pattern in VALID_BAYER:
        gc = bayer_green_compact(frame, pattern)
        return gc if gc is not None else frame
    if "RGB" in pf and frame.ndim == 3 and frame.shape[2] >= 3:
        return frame[..., 1] if (pattern in VALID_BAYER) else np.mean(frame[..., :3], axis=-1)
    if frame.ndim == 2:
        return frame
    if frame.ndim == 3 and frame.shape[2] >= 1:
        return frame[..., 1] if (pattern in VALID_BAYER and frame.shape[2] >= 2) else np.mean(frame, axis=-1)
    return frame


# -------------------------------
# Compact Style (Proxy)
# -------------------------------
class CompactProxyStyle(QProxyStyle):
    def pixelMetric(self, metric, option=None, widget=None):
        if metric in (QStyle.PM_DefaultFrameWidth, QStyle.PM_LayoutLeftMargin,
                      QStyle.PM_LayoutTopMargin, QStyle.PM_LayoutRightMargin,
                      QStyle.PM_LayoutBottomMargin, QStyle.PM_LayoutHorizontalSpacing,
                      QStyle.PM_LayoutVerticalSpacing):
            return max(2, super().pixelMetric(metric, option, widget) - 3)
        if metric in (QStyle.PM_SmallIconSize, QStyle.PM_ButtonIconSize):
            return 16
        if metric == QStyle.PM_ButtonMargin:
            return 6
        return super().pixelMetric(metric, option, widget)


# -------------------------------
# Themes (Light/Dark) — Compact + Header Dense
# -------------------------------
def _apply_light(app: QApplication):
    base_bg, card_bg = QColor("#F7F8FB"), QColor("#FFFFFF")
    text, sub = QColor("#0f172a"), QColor("#6b7280")
    accent, accent2 = QColor("#2563eb"), QColor("#1d4ed8")
    border, sel = QColor("#e5e7eb"), QColor("#e8f0fe")

    pal = QPalette()
    pal.setColor(QPalette.Window, base_bg)
    pal.setColor(QPalette.Base, card_bg)
    pal.setColor(QPalette.Text, text)
    pal.setColor(QPalette.Button, card_bg)
    pal.setColor(QPalette.ButtonText, text)
    pal.setColor(QPalette.WindowText, text)
    pal.setColor(QPalette.Highlight, sel)
    pal.setColor(QPalette.HighlightedText, text)
    app.setPalette(pal)
    app.setFont(QFont("Segoe UI", 9))  # compact font size

    app.setStyleSheet(f"""
        QWidget {{ background:{base_bg.name()}; color:{text.name()}; }}

        #HeaderBar, QGroupBox, QTableWidget, QStatusBar {{
            background:{card_bg.name()};
            border:1px solid {border.name()};
            border-radius:{RADIUS}px;
        }}
        QGroupBox {{ margin-top:14px; padding:10px 10px 8px; }}
        QGroupBox::title {{ left:10px; padding:2px 6px; font-weight:600; color:{sub.name()}; }}

        QLabel[role="title"] {{ font-weight:700; font-size:13px; color:{text.name()}; }}
        QLabel[role="badge"] {{
            background:#F3F4F6; border:1px solid {border.name()};
            border-radius:999px; padding:2px 8px; color:#334155; font-weight:600; font-size:12px;
        }}
        QLabel[status="ok"]  {{ color:#0E9F6E; background:#ECFDF5; border-color:#A7F3D0; }}
        QLabel[status="err"] {{ color:#DC2626; background:#FEF2F2; border-color:#FECACA; }}

        QLineEdit, QComboBox, QSpinBox {{
            background:{card_bg.name()}; border:1px solid {border.name()};
            border-radius:{RADIUS-2}px; padding:4px 8px; min-height:{CONTROL_H}px;
        }}
        QLineEdit:focus, QComboBox:focus, QSpinBox:focus {{ border:1px solid {accent.name()}; }}

        QPushButton {{
            background:{card_bg.name()}; color:{text.name()}; border:1px solid {border.name()};
            border-radius:{RADIUS-2}px; padding:4px 10px; min-height:{CONTROL_H}px; font-weight:600;
        }}
        QPushButton:hover {{ background:#F9FAFB; }}
        QPushButton:pressed {{ background:#F3F4F6; }}
        QPushButton[accent="true"] {{ background:{accent.name()}; border-color:{accent.name()}; color:white; }}
        QPushButton[accent="true"]:hover {{ background:{accent2.name()}; }}

        QTableWidget {{
            gridline-color:{border.name()};
            selection-background-color:{sel.name()};
            selection-color:{text.name()};
            alternate-background-color:#FAFBFF;
        }}
        QTableWidget::item {{ padding:4px; }}
        QTableView::item:hover {{ background:#F6FAFF; }}
        QHeaderView::section {{
            background:#F3F4F6; color:{sub.name()};
            padding:6px 8px; border:none; border-bottom:1px solid {border.name()};
            font-weight:600; font-size:12px;
        }}

        QProgressBar {{
            background:{card_bg.name()}; border:1px solid {border.name()};
            border-radius:999px; text-align:center; padding:1px; height:14px;
        }}
        QProgressBar::chunk {{ background:{accent.name()}; border-radius:999px; }}

        QGraphicsView {{ background:#0B1020; border:1px solid {border.name()}; border-radius:{RADIUS}px; }}
        QStatusBar {{ padding:2px 6px; }}

        /* ==== Header (Dense) =================================================== */
        #HeaderBar {{ border-radius:6px; }}
        #HeaderBar QLabel[role="title"] {{ font-weight:700; font-size:12px; }}
        #HeaderBar QLabel[role="badge"] {{ padding:1px 6px; font-size:11px; }}
        #HeaderBar QPushButton {{
            min-height:24px;  /* 헤더만 슬림 */
            padding:2px 8px;
            border-radius:6px;
        }}
    """)

def _apply_dark(app: QApplication):
    base_bg, card_bg = QColor("#0f172a"), QColor("#101826")
    text, sub = QColor("#e5e7eb"), QColor("#94a3b8")
    accent, accent2 = QColor("#3b82f6"), QColor("#2563eb")
    border, sel = QColor("#1f2937"), QColor("#1e293b")

    pal = QPalette()
    pal.setColor(QPalette.Window, base_bg)
    pal.setColor(QPalette.Base, card_bg)
    pal.setColor(QPalette.Text, text)
    pal.setColor(QPalette.Button, card_bg)
    pal.setColor(QPalette.ButtonText, text)
    pal.setColor(QPalette.WindowText, text)
    pal.setColor(QPalette.Highlight, sel)
    pal.setColor(QPalette.HighlightedText, text)
    app.setPalette(pal)
    app.setFont(QFont("Segoe UI", 9))

    app.setStyleSheet(f"""
        QWidget {{ background:{base_bg.name()}; color:{text.name()}; }}

        #HeaderBar, QGroupBox, QTableWidget, QStatusBar {{
            background:{card_bg.name()};
            border:1px solid {border.name()};
            border-radius:{RADIUS}px;
        }}
        QGroupBox {{ margin-top:14px; padding:10px 10px 8px; }}
        QGroupBox::title {{ left:10px; padding:2px 6px; font-weight:600; color:{sub.name()}; }}

        QLabel[role="title"] {{ font-weight:700; font-size:13px; color:#bfdbfe; }}
        QLabel[role="badge"] {{
            background:#0b1220; border:1px solid {border.name()};
            border-radius:999px; padding:2px 8px; color:#9db7d7; font-weight:600; font-size:12px;
        }}
        QLabel[status="ok"]  {{ color:#34d399; background:#064e3b; border-color:#065f46; }}
        QLabel[status="err"] {{ color:#fca5a5; background:#7f1d1d; border-color:#991b1b; }}

        QLineEdit, QComboBox, QSpinBox {{
            background:#0b1220; border:1px solid {border.name()};
            border-radius:{RADIUS-2}px; padding:4px 8px; min-height:{CONTROL_H}px; color:{text.name()};
        }}
        QLineEdit:focus, QComboBox:focus, QSpinBox:focus {{ border:1px solid {accent.name()}; }}

        QPushButton {{
            background:#0b1220; color:{text.name()}; border:1px solid {border.name()};
            border-radius:{RADIUS-2}px; padding:4px 10px; min-height:{CONTROL_H}px; font-weight:600;
        }}
        QPushButton:hover {{ background:#101826; }}
        QPushButton:pressed {{ background:#0c1320; }}
        QPushButton[accent="true"] {{ background:{accent.name()}; border-color:{accent.name()}; color:white; }}
        QPushButton[accent="true"]:hover {{ background:{accent2.name()}; }}

        QTableWidget {{
            gridline-color:{border.name()};
            selection-background-color:{sel.name()};
            selection-color:{text.name()};
            alternate-background-color:#0e1621;
        }}
        QTableWidget::item {{ padding:4px; }}
        QTableView::item:hover {{ background:#101a2a; }}
        QHeaderView::section {{
            background:#111827; color:{sub.name()};
            padding:6px 8px; border:none; border-bottom:1px solid {border.name()};
            font-weight:600; font-size:12px;
        }}

        QProgressBar {{
            background:{card_bg.name()}; border:1px solid {border.name()};
            border-radius:999px; text-align:center; padding:1px; height:14px;
        }}
        QProgressBar::chunk {{ background:{accent.name()}; border-radius:999px; }}

        QGraphicsView {{ background:#0B1020; border:1px solid {border.name()}; border-radius:{RADIUS}px; }}
        QStatusBar {{ padding:2px 6px; }}

        /* ==== Header (Dense) =================================================== */
        #HeaderBar {{ border-radius:6px; }}
        #HeaderBar QLabel[role="title"] {{ font-weight:700; font-size:12px; color:#bfdbfe; }}
        #HeaderBar QLabel[role="badge"] {{ padding:1px 6px; font-size:11px; }}
        #HeaderBar QPushButton {{
            min-height:24px;
            padding:2px 8px;
            border-radius:6px;
        }}
    """)


# -------------------------------
# Main Window
# -------------------------------
class EMVAAnalyzerApp(QMainWindow):
    def __init__(self, forced_icon_path: Optional[str] = None, show_header: bool = True):
        super().__init__()
        self.setWindowTitle("EMVA 1288 Analyzer")
        self.resize(1240, 800)  # 컴팩트 기본 크기

        # 헤더 표시 여부
        self._show_header = show_header
        # 헤더 위젯이 없을 수도 있으므로 미리 None으로 초기화
        self.header_status = None
        self.quick_apply_btn = None
        self.quick_live_btn = None

        # 강제 아이콘 경로(있으면 WinAPI로도 적용)
        self._icon_path_forced = forced_icon_path

        # 상태
        self.camera: Optional[CxpCamera] = None
        self.lightbox: Optional[LightBoxController] = None
        self.measurement_worker: Optional[MeasurementWorker] = None
        self.dark_data_worker: Optional[DarkDataWorker] = None
        self.sat_worker: Optional[SaturationWorker] = None
        self.linearity_results: List[Dict] = []
        self.dark_data_results: List[Dict] = []
        self.live_timer = QTimer(self)
        self.live_timer.timeout.connect(self.update_live_view)
        self.debug = True
        self.is_line_camera = False
        self.table_mode = "idle"
        self.lb_max_units = DEFAULT_LB_MAX_UNITS
        self.bayer_pattern_default = "None"

        self.initUI()
        self.connect_devices()

        # Show 후 작업표시줄/Alt-Tab 아이콘까지 WinAPI로 강제
        if sys.platform.startswith("win") and self._icon_path_forced:
            QTimer.singleShot(120, self._force_taskbar_icon_once)


    def _force_taskbar_icon_once(self):
        try:
            hwnd = int(self.winId().__int__())
            _force_window_icon_win(hwnd, self._icon_path_forced)
        except Exception as e:
            print(f"[ICON-FORCE] failed: {e}")

    # ---------------- UI ----------------
    def _header_bar(self) -> QWidget:
        bar = QWidget()
        bar.setObjectName("HeaderBar")
        h = QHBoxLayout(bar)
        h.setContentsMargins(3, 2, 3, 2)   # 더 슬림
        h.setSpacing(3)

        title = QLabel("EMVA 1288 Analyzer")
        title.setProperty("role", "title")

        self.header_status = QLabel("Disconnected")
        self.header_status.setProperty("role", "badge")
        self.header_status.setProperty("status", "err")

        h.addWidget(title)
        h.addSpacing(6)
        h.addWidget(self.header_status)
        h.addStretch(1)

        self.quick_apply_btn = QPushButton("Apply")
        self.quick_apply_btn.setProperty("accent", True)
        self.quick_apply_btn.setIcon(self.style().standardIcon(QStyle.SP_DialogApplyButton))
        self.quick_apply_btn.clicked.connect(self.apply_camera_parameters)
        self.quick_apply_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.quick_apply_btn.setIconSize(QSize(14, 14))
        self.quick_apply_btn.setMinimumHeight(24)

        self.quick_live_btn = QPushButton("Live")
        self.quick_live_btn.setIcon(self.style().standardIcon(QStyle.SP_MediaPlay))
        self.quick_live_btn.setCheckable(True)
        self.quick_live_btn.clicked.connect(lambda c: self.toggle_live_view(c))
        self.quick_live_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.quick_live_btn.setIconSize(QSize(14, 14))
        self.quick_live_btn.setMinimumHeight(24)

        h.addWidget(self.quick_apply_btn, 0, Qt.AlignRight)
        h.addWidget(self.quick_live_btn, 0, Qt.AlignRight)
        return bar

    def initUI(self):
        main = QWidget()
        self.setCentralWidget(main)
        outer = QVBoxLayout(main)
        outer.setContentsMargins(8, 8, 8, 8)  # 살짝 더 슬림
        outer.setSpacing(6)

        # 옵션에 따라 헤더바 생성/생략
        if self._show_header:
            header = self._header_bar()
            outer.addWidget(header)

        splitter = QSplitter(Qt.Horizontal)
        splitter.setHandleWidth(6)
        splitter.setCollapsible(0, False)
        splitter.setCollapsible(1, False)

        # 좌측(스크롤러)
        left_content = QWidget()
        left_layout = QVBoxLayout(left_content)
        left_layout.setSpacing(8)

        # Connection
        conn_group = QGroupBox("Connection")
        conn_layout = QHBoxLayout()
        self.conn_status_label = QLabel("Status: Disconnected")
        self.conn_status_label.setProperty("role", "badge")
        self.conn_status_label.setProperty("status", "err")
        self.connect_btn = QPushButton("Reconnect")
        self.connect_btn.setIcon(self.style().standardIcon(QStyle.SP_BrowserReload))
        self.connect_btn.clicked.connect(self.connect_devices)
        self.connect_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        conn_layout.addWidget(self.conn_status_label, 1)
        conn_layout.addWidget(self.connect_btn, 0, Qt.AlignRight)
        conn_group.setLayout(conn_layout)
        left_layout.addWidget(conn_group)

        # Camera Settings
        param_group = QGroupBox("Camera Settings")
        param_layout = QFormLayout()
        param_layout.setLabelAlignment(Qt.AlignRight)
        param_layout.setHorizontalSpacing(8)
        param_layout.setVerticalSpacing(6)
        self.pixelformat_combo = QComboBox()
        self.exposure_label = QLabel("Exposure (us):")
        self.exposure_edit = QLineEdit("10000.0")
        self.blacklevel_edit = QLineEdit("0.0")
        self.bayer_pattern_combo = QComboBox()
        self.bayer_pattern_combo.addItems(["None"] + VALID_BAYER)
        self.bayer_pattern_combo.setCurrentText(self.bayer_pattern_default)
        param_layout.addRow("PixelFormat:", self.pixelformat_combo)
        param_layout.addRow(self.exposure_label, self.exposure_edit)
        param_layout.addRow("Black Level:", self.blacklevel_edit)
        param_layout.addRow("Bayer Pattern:", self.bayer_pattern_combo)
        param_group.setLayout(param_layout)
        left_layout.addWidget(param_group)

        self.apply_params_btn = QPushButton("Apply")
        self.apply_params_btn.setProperty("accent", True)
        self.apply_params_btn.setIcon(self.style().standardIcon(QStyle.SP_DialogApplyButton))
        self.apply_params_btn.clicked.connect(self.apply_camera_parameters)
        self.apply_params_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        left_layout.addWidget(self.apply_params_btn)

        # Live View
        live_group = QGroupBox("Live View")
        live_layout = QHBoxLayout()
        self.live_view_btn = QPushButton("Start Live")
        self.live_view_btn.setCheckable(True)
        self.live_view_btn.setIcon(self.style().standardIcon(QStyle.SP_MediaPlay))
        self.live_view_btn.clicked.connect(self.toggle_live_view)
        self.live_view_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        live_layout.addWidget(self.live_view_btn, 0, Qt.AlignLeft)
        live_group.setLayout(live_layout)
        left_layout.addWidget(live_group)

        # EMVA (Linearity)
        emva_group = QGroupBox("EMVA 1288 (Linearity)")
        emva_layout = QVBoxLayout()
        self.num_steps_spin = QSpinBox()
        self.num_steps_spin.setRange(10, 200)
        self.num_steps_spin.setValue(20)
        self.num_steps_spin.setPrefix("Steps: ")
        self.start_full_measurement_btn = QPushButton("Start Linearity")
        self.start_full_measurement_btn.setIcon(self.style().standardIcon(QStyle.SP_ArrowRight))
        self.start_full_measurement_btn.clicked.connect(self.start_full_measurement)
        self.start_full_measurement_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.stop_measurement_btn = QPushButton("Stop")
        self.stop_measurement_btn.setIcon(self.style().standardIcon(QStyle.SP_MediaStop))
        self.stop_measurement_btn.clicked.connect(self.stop_measurement)
        self.stop_measurement_btn.setEnabled(False)
        self.stop_measurement_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        measure_btn_layout = QHBoxLayout()
        measure_btn_layout.addWidget(self.start_full_measurement_btn)
        measure_btn_layout.addWidget(self.stop_measurement_btn)
        emva_layout.addWidget(self.num_steps_spin)
        emva_layout.addLayout(measure_btn_layout)
        emva_group.setLayout(emva_layout)
        left_layout.addWidget(emva_group)

        # Dark Current
        dark_group = QGroupBox("Dark Current Measurement")
        dark_layout = QFormLayout()
        dark_layout.setLabelAlignment(Qt.AlignRight)
        dark_layout.setHorizontalSpacing(8)
        dark_layout.setVerticalSpacing(6)
        self.dark_exp_start_edit = QLineEdit("0.10")
        self.dark_exp_end_edit = QLineEdit("1.00")
        self.dark_exp_steps_edit = QLineEdit("20")
        self.dark_frames_edit = QLineEdit("16")
        self.system_gain_edit = QLineEdit("")
        self.dark_threshold_edit = QLineEdit("")
        self.dark_threshold_edit.setPlaceholderText("e.g. 3800 (DN). Empty = no threshold")
        self.start_dark_data_btn = QPushButton("Start Dark")
        self.start_dark_data_btn.setIcon(self.style().standardIcon(QStyle.SP_ArrowRight))
        self.start_dark_data_btn.clicked.connect(self.start_dark_data_measurement)
        self.start_dark_data_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        dark_layout.addRow("Exp Start (s):", self.dark_exp_start_edit)
        dark_layout.addRow("Exp End (s):", self.dark_exp_end_edit)
        dark_layout.addRow("Num Steps:", self.dark_exp_steps_edit)
        dark_layout.addRow("Frames per step:", self.dark_frames_edit)
        dark_layout.addRow("Exclude pixels > threshold (DN):", self.dark_threshold_edit)
        dark_layout.addRow(self.start_dark_data_btn)
        dark_group.setLayout(dark_layout)
        left_layout.addWidget(dark_group)

        # Export
        export_group = QGroupBox("Export")
        ex_layout = QVBoxLayout()
        self.export_btn = QPushButton("Export Excel")
        self.export_btn.setIcon(self.style().standardIcon(QStyle.SP_DriveHDIcon))
        self.export_btn.clicked.connect(self.export_all_to_excel)
        self.export_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        ex_layout.addWidget(self.export_btn)
        export_group.setLayout(ex_layout)
        left_layout.addWidget(export_group)

        left_layout.addStretch(1)
        self.status_label_bottom = QLabel("Ready")
        left_layout.addWidget(self.status_label_bottom)

        # 좌측 스크롤 컨테이너
        left_scroll = QScrollArea()
        left_scroll.setWidgetResizable(True)
        left_scroll.setFrameShape(QFrame.NoFrame)
        left_scroll.setWidget(left_content)

        # 우측(디스플레이)
        right = QWidget()
        rightv = QVBoxLayout(right)
        rightv.setSpacing(8)

        title_live = QLabel("Live Image")
        title_live.setProperty("role", "title")
        rightv.addWidget(title_live, 0, Qt.AlignLeft)

        self.graphics_view = QGraphicsView()
        self.scene = QGraphicsScene()
        self.graphics_view.setScene(self.scene)
        self.graphics_view.setRenderHint(QPainter.Antialiasing, True)
        self.graphics_view.setRenderHint(QPainter.SmoothPixmapTransform, True)
        self.graphics_view.setMinimumHeight(420)
        rightv.addWidget(self.graphics_view, 2)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        rightv.addWidget(sep)

        title_tbl = QLabel("Measurement Results")
        title_tbl.setProperty("role", "title")
        rightv.addWidget(title_tbl, 0, Qt.AlignLeft)

        self.result_table = QTableWidget()
        self.set_result_table_headers_for("linearity")
        self.result_table.setAlternatingRowColors(True)
        self.result_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.result_table.setSelectionMode(QTableWidget.SingleSelection)
        self.result_table.verticalHeader().setVisible(False)
        self.result_table.horizontalHeader().setStretchLastSection(True)
        self.result_table.setMinimumHeight(220)
        self.result_table.setShowGrid(False)
        self.result_table.setWordWrap(False)
        self.result_table.horizontalHeader().setDefaultSectionSize(120)
        rightv.addWidget(self.result_table, 1)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        rightv.addWidget(self.progress_bar)

        # Splitter 구성
        splitter.addWidget(left_scroll)
        splitter.addWidget(right)
        splitter.setSizes([520, 900])
        outer.addWidget(splitter)

        # StatusBar
        self.setStatusBar(QStatusBar())
        self.statusBar().showMessage("Ready")

    # ----- 테이블 헤더 -----
    def set_result_table_headers_for(self, mode: str):
        self.table_mode = mode
        self.result_table.clear()
        if mode == "linearity":
            self.result_table.setColumnCount(4)
            self.result_table.setHorizontalHeaderLabels(
                ["Light Level", "Gray Value", "Temporal STD", "Total STD"]
            )
        elif mode == "dark":
            # ⬇️ Total Noise 열 추가
            self.result_table.setColumnCount(4)
            self.result_table.setHorizontalHeaderLabels(
                ["Exposure (s)", "Mean Gray (DN)", "Temporal Noise (DN)", "Total Noise (DN)"]
            )
        else:
            self.result_table.setColumnCount(0)
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.result_table.setRowCount(0)

    # ---------------- 라이브뷰/표시 ----------------
    def _reshape_if_rgb_unpacked(self, frame: np.ndarray) -> Optional[np.ndarray]:
        if frame is None or not self.camera:
            return None
        pf_up = (self.pixelformat_combo.currentText() or "").upper()
        is_rgb = ("RGB" in pf_up)
        if is_rgb and (frame.ndim != 3 or frame.shape[2] != 3):
            try:
                h = int(self.camera.get("Height"))
                w = int(self.camera.get("Width"))
                if frame.dtype != np.uint16:
                    frame = frame.view(np.uint16)
                return frame.reshape((h, w, 3))
            except Exception as e:
                print(f"reshape RGB failed: {e}")
                return None
        return frame

    def display_image(self, frame: np.ndarray):
        if frame is None or frame.size == 0:
            return
        pixel_format_str = self.pixelformat_combo.currentText()
        selected = self.bayer_pattern_combo.currentText()
        pattern = selected if selected in VALID_BAYER else None
        gmono = to_g_mono(frame, pixel_format_str, pattern)
        if gmono is None:
            return
        if pattern and gmono.ndim == 2:
            gshow = upsample_nn2x(gmono)
        else:
            gshow = gmono if gmono.ndim == 2 else (
                gmono[..., 1] if (gmono.ndim == 3 and gmono.shape[2] >= 2) else np.mean(gmono, axis=-1)
            )
        pf_up = (pixel_format_str or "").upper()
        scale = 16 if "12" in pf_up else (4 if "10" in pf_up else 256)
        disp8 = np.clip(gshow / scale, 0, 255).astype(np.uint8)
        qimg = QImage(disp8.data, gshow.shape[1], gshow.shape[0], gshow.shape[1], QImage.Format_Grayscale8).copy()
        self.scene.clear()
        self.scene.addPixmap(QPixmap.fromImage(qimg))
        self.graphics_view.fitInView(self.scene.itemsBoundingRect(), Qt.KeepAspectRatio)
        mode_text = f"PF={pixel_format_str}, Pattern={selected}, Mode={'G-only(compact+upsample)' if pattern else 'Original'}"
        self.status_label_bottom.setText("Live • updating")
        self.statusBar().showMessage(f"Live view… [{mode_text}]")

    @pyqtSlot()
    def update_live_view(self):
        try:
            raw_frame, _ = self.camera.snap_pair(delay_ms=1)
            f = self._reshape_if_rgb_unpacked(raw_frame)
            if f is not None:
                self.display_image(f)
        except Exception as e:
            print(f"live view error: {e}")
            self.toggle_live_view(False)

    @pyqtSlot(object)
    def on_sat_preview(self, frame):
        try:
            f = self._reshape_if_rgb_unpacked(frame)
            if f is not None:
                self.display_image(f)
        except Exception as e:
            print(f"saturation preview error: {e}")

    # ---------------- 상태/측정 플로우 ----------------
    def is_any_task_running(self) -> bool:
        return any([
            self.measurement_worker and self.measurement_worker.isRunning(),
            self.dark_data_worker and self.dark_data_worker.isRunning(),
            self.sat_worker and self.sat_worker.isRunning()
        ])

    def start_full_measurement(self):
        if not self.camera or not self.lightbox:
            return self.show_error_message("Devices are not connected.")
        if self.is_any_task_running():
            return self.show_error_message("A task is already in progress.")
        if self.live_view_btn.isChecked():
            self.toggle_live_view(False)
            time.sleep(0.05)

        # 테이블/상태 초기화
        self.result_table.setRowCount(0)
        self.linearity_results.clear()
        self.set_result_table_headers_for("linearity")
        self.set_ui_for_measurement(True)
        self.status_label_bottom.setText("Measuring…")
        self.statusBar().showMessage("Step 1: Finding saturation level…")
        QApplication.processEvents()

        selected_pattern = (
            self.bayer_pattern_combo.currentText()
            if self.bayer_pattern_combo.currentText() in VALID_BAYER
            else None
        )

        # 1) 포화 지점 탐색 워커 실행 (기존 로직 그대로)
        self.sat_worker = SaturationWorker(
            camera=self.camera,
            lightbox=self.lightbox,
            pixel_format=self.pixelformat_combo.currentText(),
            bayer_pattern=selected_pattern,
            lb_max_units=int(self.lb_max_units),
        )
        self.sat_worker.progress_updated.connect(self.update_progress)
        self.sat_worker.preview_ready.connect(self.on_sat_preview)

        loop = QEventLoop()
        first_clip_holder = {"val": None}

        def _on_sat(v):
            first_clip_holder["val"] = v

        def _on_err(m):
            self.show_error_message(f"Saturation search error: {m}")

        self.sat_worker.saturation_found.connect(_on_sat)
        self.sat_worker.error_occurred.connect(_on_err)
        self.sat_worker.finished.connect(loop.quit)
        self.sat_worker.start()
        loop.exec_()
        self.sat_worker = None

        first_clip = first_clip_holder["val"]
        if not first_clip:
            self.set_ui_for_measurement(False)
            return

        # 2) 선형 구간 상한을 "포화 직전"으로 낮추기
        n = int(self.num_steps_spin.value())

        # 안전 계수: 찾아낸 first_clip 의 몇 % 까지만 선형으로 쓸지
        #  0.90 이면 90%까지만 사용
        SAFETY_RATIO = 0.90

        max_linear_level = int(first_clip * SAFETY_RATIO)

        # 최소/최대 범위 보정
        max_linear_level = max(1, max_linear_level)
        max_linear_level = min(max_linear_level, int(self.lb_max_units))

        print(
            f"[LinearityInit] first_clip={first_clip}, "
            f"max_linear_level={max_linear_level}, "
            f"safety_ratio={SAFETY_RATIO}, steps={n}"
        )

        self.statusBar().showMessage(
            f"First-clip={first_clip} → using up to {max_linear_level} "
            f"(≈{SAFETY_RATIO * 100:.0f}% of clip) for {n} linear steps…"
        )
        QApplication.processEvents()

        # 3) 실제 선형 스텝 생성: 0 ~ max_linear_level 사이를 균일 분할
        linear_levels = compute_light_steps(max_linear_level, n)

        # 3-1) 포화 근처 검증 스텝 2~3 개 추가
        VERIFY_STEPS = 3
        extra_levels = []

        if first_clip > max_linear_level:
            span = first_clip - max_linear_level
            # 선형 구간 상단 ~ 포화 지점 사이를 균등 분할해서 찍어보기
            step = max(1, span // (VERIFY_STEPS + 1))
            for k in range(1, VERIFY_STEPS + 1):
                lv = max_linear_level + k * step
                if lv <= self.lb_max_units:
                    extra_levels.append(int(lv))

            # first_clip 자체가 리스트에 없으면 포함
            if first_clip <= self.lb_max_units and first_clip not in extra_levels:
                extra_levels.append(int(first_clip))

        # 전체 레벨: 선형 구간 + 검증 구간
        all_levels = sorted(set(linear_levels + extra_levels))

        # 선형 구간의 마지막 인덱스 (all_levels 기준)
        # compute_light_steps 가 0~max_linear_level 까지 올라가므로
        # 정렬하면 항상 선형 레벨들이 앞에, 그 뒤에 extra_levels 가 온다고 가정
        pre_last_index = len(linear_levels) - 1

        print(
            f"[LinearityInit] linear_levels={len(linear_levels)}, "
            f"extra_levels={extra_levels}, all_levels={len(all_levels)}, "
            f"pre_last_index={pre_last_index}"
        )

        # 4) 라인성 측정 워커 시작
        self.measurement_worker = MeasurementWorker(
            self.camera,
            self.lightbox,
            all_levels,
            pixel_format=self.pixelformat_combo.currentText(),
            bayer_pattern=selected_pattern,
            lb_max_units=int(self.lb_max_units),
            pre_last_index=pre_last_index,
        )
        self.measurement_worker.progress_updated.connect(self.update_progress)
        self.measurement_worker.result_ready.connect(self.add_result_to_table)
        self.measurement_worker.finished.connect(self.measurement_finished)
        self.measurement_worker.error_occurred.connect(self.on_measurement_error)
        self.measurement_worker.start()


    def start_dark_data_measurement(self):
        if not self.camera:
            return self.show_error_message("Camera is not connected.")
        if self.is_any_task_running():
            return self.show_error_message("Another task is already in progress.")
        if self.live_view_btn.isChecked():
            self.toggle_live_view(False)
            time.sleep(0.05)

        if self.lightbox is not None:
            with suppress(Exception):
                # 밝기 0으로
                self.lightbox.set_light_level(0)
            with suppress(Exception):
                # 완전히 끄고 싶으면 전원 OFF
                self.lightbox.power_control(turn_on=False)
        try:
            exp_start_s = float(self.dark_exp_start_edit.text())
            exp_end_s = float(self.dark_exp_end_edit.text())
            num_steps = int(self.dark_exp_steps_edit.text())
            frames_per_step = int(self.dark_frames_edit.text() or "16")
            if frames_per_step < 2:
                raise ValueError("Frames per step must be >= 2 for pairwise (EMVA).")
            threshold_dn = None
            t = (self.dark_threshold_edit.text() or "").strip()
            if t:
                threshold_dn = float(t)
                if threshold_dn < 0:
                    raise ValueError("Threshold must be >= 0 DN.")
            if exp_start_s >= exp_end_s or num_steps < 2:
                raise ValueError("Invalid range/steps.")
        except Exception as e:
            return self.show_error_message(f"Invalid input: {e}")

        self.result_table.setRowCount(0)
        self.set_result_table_headers_for("dark")
        self.dark_data_results.clear()
        self.set_ui_for_measurement(True)
        self.status_label_bottom.setText("Measuring dark…")
        self.statusBar().showMessage("Starting Dark & Read-Noise (pairwise)…")
        QApplication.processEvents()

        exposure_steps = np.linspace(exp_start_s * 1_000_000.0, exp_end_s * 1_000_000.0, num_steps).tolist()
        self.dark_data_worker = DarkDataWorker(
            camera=self.camera,
            exposure_steps=exposure_steps,
            pixel_format=self.pixelformat_combo.currentText(),
            frames_to_average=int(self.dark_frames_edit.text()),
            system_gain_e_per_dn=None,
            compute_per_pixel_pairwise=False,
            threshold_dn=threshold_dn
        )
        self.dark_data_worker.progress_updated.connect(self.update_progress)
        self.dark_data_worker.result_ready.connect(self.add_dark_data_result)
        self.dark_data_worker.finished.connect(self.measurement_finished)
        self.dark_data_worker.error_occurred.connect(self.on_measurement_error)
        self.dark_data_worker.start()

    @pyqtSlot(dict)
    def add_result_to_table(self, r: dict):
        self.linearity_results.append({
            "light_level": float(r["light_level"]),
            "gray_value_subtracted": float(r["gray_value_subtracted"]),
            "temporal_std": float(r["temporal_std"]),
            "total_std": float(r["total_std"]),
        })
        if self.table_mode != "linearity":
            self.set_result_table_headers_for("linearity")
        row = self.result_table.rowCount()
        self.result_table.insertRow(row)

        def _num_item(v: float, fmt: str = ".2f"):
            it = QTableWidgetItem(f"{v:{fmt}}")
            it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            return it

        self.result_table.setItem(row, 0, _num_item(r['light_level']))
        gv_disp = float(r.get("gray_value_display", r["gray_value_subtracted"]))
        self.result_table.setItem(row, 1, _num_item(gv_disp))
        self.result_table.setItem(row, 2, _num_item(r['temporal_std']))
        self.result_table.setItem(row, 3, _num_item(r['total_std']))
        is_plateau = r.get("is_plateau", False)
        self.status_label_bottom.setText(f"Linearity • {row+1}")
        self.statusBar().showMessage(f"Linearity step {row+1}{' (Plateau)' if is_plateau else ''}")
        self.display_image(r['last_frame'])

    @pyqtSlot(dict)
    def add_dark_data_result(self, r: dict):
        exp_s = float(r.get("exposure_time_s",
                            float(r.get("exposure_time_us", r.get("exposure_time", 0.0))) / 1_000_000.0))
        temporal_dn = r.get("read_noise_dn")
        total_dn = r.get("total_noise_dn")
        self.dark_data_results.append({
            "exposure_time": exp_s,
            "mean_gray_value": float(r.get("mean_gray_value", 0.0)),
            "temporal_noise_dn": (float(temporal_dn) if temporal_dn is not None else None),
            "total_noise_dn": (float(total_dn) if total_dn is not None else None),
            "method": r.get("method", "pairwise"),
        })
        if self.table_mode != "dark":
            self.set_result_table_headers_for("dark")
        row = self.result_table.rowCount()
        self.result_table.insertRow(row)

        # ⬇️ 기본 포맷을 소수 6째 자리까지로 변경
        def _num_item(v: float, fmt: str = ".6f"):
            it = QTableWidgetItem(f"{v:{fmt}}")
            it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            return it

        # Exposure (s)  → 6자리
        self.result_table.setItem(row, 0, _num_item(exp_s))

        # Mean Gray (DN) → 그대로 6자리
        self.result_table.setItem(row, 1, _num_item(float(r.get('mean_gray_value', 0.0)), ".6f"))

        # Temporal Noise (DN) → 4자리 정도로 확장
        if temporal_dn is None:
            it = QTableWidgetItem("N/A")
            it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.result_table.setItem(row, 2, it)
        else:
            self.result_table.setItem(row, 2, _num_item(float(temporal_dn), ".4f"))

        # Total Noise (DN) → 4자리 정도로 확장
        if total_dn is None:
            it = QTableWidgetItem("N/A")
            it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.result_table.setItem(row, 3, it)
        else:
            self.result_table.setItem(row, 3, _num_item(float(total_dn), ".4f"))

        self.result_table.scrollToBottom()


    def stop_measurement(self):
        if self.measurement_worker and self.measurement_worker.isRunning():
            self.measurement_worker.stop()
        if self.dark_data_worker and self.dark_data_worker.isRunning():
            self.dark_data_worker.stop()
        if self.sat_worker and self.sat_worker.isRunning():
            self.sat_worker.stop()
        self.set_ui_for_measurement(False)
        self.statusBar().showMessage("Stopped.")

    def set_ui_for_measurement(self, running: bool):
        widgets = [
            self.start_full_measurement_btn, self.start_dark_data_btn, self.stop_measurement_btn,
            self.connect_btn, self.apply_params_btn, self.live_view_btn, self.export_btn,
            self.quick_apply_btn, self.quick_live_btn
        ]
        for w in (w for w in widgets if w is not None):
            if w is self.stop_measurement_btn:
                w.setEnabled(running)
            else:
                w.setEnabled(not running)

    @pyqtSlot(int, str)
    def update_progress(self, v, msg):
        self.progress_bar.setValue(v)
        self.status_label_bottom.setText(msg if len(msg) < 40 else msg[:38] + "…")
        self.statusBar().showMessage(msg)

    @pyqtSlot()
    def measurement_finished(self):
        self.status_label_bottom.setText("Measurement finished.")
        self.statusBar().showMessage("Measurement finished.")
        self.set_ui_for_measurement(False)
        self.measurement_worker = None
        self.dark_data_worker = None
        self.sat_worker = None

    @pyqtSlot(str)
    def on_measurement_error(self, m):
        self.show_error_message(f"Measurement Error: {m}")
        self.measurement_finished()

    def export_all_to_excel(self):
        """
        Linearity/Dark 데이터를 템플릿 복제본에 openpyxl로 안전하게 기록.
        - 템플릿은 실행 컨텍스트 기준 절대경로 탐색(_template_path)
        - 데이터 구간과 겹치는 병합셀은 먼저 해제
        - A5부터 4열에 기록 (템플릿 헤더/수식 보존)
        - UI 리스트(self.linearity_results / self.dark_data_results) 그대로 사용
        """
        import openpyxl
        from shutil import copy2
        from pathlib import Path
        import os, stat

        # 1) 데이터 유무 확인
        has_lin = len(self.linearity_results) > 0
        has_dark = len(self.dark_data_results) > 0
        if not (has_lin or has_dark):
            return self.show_error_message("No data to export. Run a measurement first.")

        save_path, _ = QFileDialog.getSaveFileName(
            self, "Save Results As", "EMVA1288_Results.xlsx", "Excel Files (*.xlsx)"
        )
        if not save_path:
            return

        # 2) 템플릿 찾기 + 복사 + 읽기전용 해제
        tpl = _template_path("Template_format.xlsx")
        if not tpl.exists():
            return self.show_error_message(f"'Template_format.xlsx' not found.\nTried: {tpl}")
        dst = Path(save_path).resolve()
        dst.parent.mkdir(parents=True, exist_ok=True)
        try:
            copy2(str(tpl), str(dst))
            # 복사본이 읽기전용이면 openpyxl.save가 실패/침묵할 수 있음 → writable로
            try:
                os.chmod(str(dst), stat.S_IWRITE | stat.S_IREAD)
            except Exception:
                pass
        except Exception as e:
            return self.show_error_message(f"Failed to copy template: {e}")

        # 3) openpyxl로 복제본 열기
        try:
            wb = openpyxl.load_workbook(str(dst))
        except Exception as e:
            return self.show_error_message(f"Failed to open copied template: {e}")

        # 시트명(템플릿과 일치)
        SHEET_LINEARITY = "Linearity Data"
        SHEET_DARK = "Linearity Data(Exposure)"

        # 데이터 구간 정의 (템플릿 규칙)
        HDR_ROW = 4  # 헤더/가이드 라인(병합 있을 수 있음)
        DATA_START_ROW = 5
        # 템플릿에 맞춰 100행(5~104) 기본. 더 쓰고 싶으면 여길 늘리거나 동적 계산 가능.
        DATA_MAX_ROWS = 100
        DATA_END_ROW = DATA_START_ROW + DATA_MAX_ROWS - 1
        MIN_COL, MAX_COL = 1, 4  # A..D

        def _ranges_intersect(rng, r1, r2, c1, c2):
            return not (rng.max_row < r1 or rng.min_row > r2 or rng.max_col < c1 or rng.min_col > c2)

        def _unmerge_intersections(ws, min_row, max_row, min_col, max_col):
            to_unmerge = [rng for rng in list(ws.merged_cells.ranges)
                          if _ranges_intersect(rng, min_row, max_row, min_col, max_col)]
            for rng in to_unmerge:
                ws.unmerge_cells(str(rng))

        def _clear_range(ws, min_row, max_row, min_col, max_col):
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    cell.value = None

        try:
            # ---------------- Linearity ----------------
            if has_lin:
                if SHEET_LINEARITY in wb.sheetnames:
                    ws_lin = wb[SHEET_LINEARITY]
                else:
                    ws_lin = wb.create_sheet(SHEET_LINEARITY)

                # 병합 해제(헤더~데이터 구간 전체)
                _unmerge_intersections(ws_lin, HDR_ROW, DATA_END_ROW, MIN_COL, MAX_COL)
                # 기존 데이터 지우기
                _clear_range(ws_lin, DATA_START_ROW, DATA_END_ROW, MIN_COL, MAX_COL)

                # 쓰기 (A: Light Level, B: Gray(sub), C: Temporal, D: Total)
                for i, r in enumerate(self.linearity_results):
                    row_idx = DATA_START_ROW + i
                    if row_idx > DATA_END_ROW:
                        break
                    ws_lin.cell(row=row_idx, column=1, value=float(r["light_level"]))
                    ws_lin.cell(row=row_idx, column=2, value=float(r["gray_value_subtracted"]))
                    ws_lin.cell(row=row_idx, column=3, value=float(r["temporal_std"]))
                    ws_lin.cell(row=row_idx, column=4, value=float(r["total_std"]))

            # ---------------- Dark ----------------
            if has_dark:
                if SHEET_DARK in wb.sheetnames:
                    ws_dark = wb[SHEET_DARK]
                else:
                    ws_dark = wb.create_sheet(SHEET_DARK)

                _unmerge_intersections(ws_dark, HDR_ROW, DATA_END_ROW, MIN_COL, MAX_COL)
                _clear_range(ws_dark, DATA_START_ROW, DATA_END_ROW, MIN_COL, MAX_COL)

                # 쓰기 (A: Exposure(s), B: Mean Gray, C: Temporal Noise, D: Total Noise)
                for i, r in enumerate(self.dark_data_results):
                    row_idx = DATA_START_ROW + i
                    if row_idx > DATA_END_ROW:
                        break
                    ws_dark.cell(row=row_idx, column=1, value=float(r.get("exposure_time", 0.0)))
                    ws_dark.cell(row=row_idx, column=2, value=float(r.get("mean_gray_value", 0.0)))
                    tn = r.get("temporal_noise_dn")
                    ws_dark.cell(row=row_idx, column=3, value=(float(tn) if tn is not None else None))
                    tot = r.get("total_noise_dn")
                    ws_dark.cell(row=row_idx, column=4, value=(float(tot) if tot is not None else None))

            # 4) 저장
            wb.save(str(dst))
            QMessageBox.information(self, "Success", f"Results successfully saved to:\n{dst}")

        except Exception as e:
            self.show_error_message(f"Failed to export to Excel: {e}\nTemplate: {str(dst)}")

    # ---------------- 공통/장치 ----------------
    def show_error_message(self, message: str):
        QMessageBox.critical(self, "Error", message)

    def closeEvent(self, e):
        self.live_timer.stop()
        self.stop_measurement()
        time.sleep(0.5)
        if self.lightbox:
            with contextlib.suppress(Exception):
                self.lightbox.close()
        if self.camera:
            with contextlib.suppress(Exception):
                self.camera.ctrl.disconnect_camera()
        e.accept()

    def connect_devices(self):
        try:
            self.status_label_bottom.setText("Connecting...")
            self.statusBar().showMessage("Connecting...")
            QApplication.processEvents()
            self.camera = CxpCamera()
            self.lightbox = LightBoxController(debug=self.debug)
            if not self.camera.connect():
                raise ConnectionError("Camera not found.")
            self.conn_status_label.setText("Status: Connected")
            self.conn_status_label.setProperty("status", "ok")
            self.conn_status_label.style().unpolish(self.conn_status_label)
            self.conn_status_label.style().polish(self.conn_status_label)

            # 헤더가 있을 때만 갱신
            if self.header_status is not None:
                self.header_status.setText("Connected")
                self.header_status.setProperty("status", "ok")
                self.header_status.style().unpolish(self.header_status)
                self.header_status.style().polish(self.header_status)

            self.status_label_bottom.setText("Devices connected successfully.")
            self.statusBar().showMessage("Devices connected successfully.")
            self.load_camera_parameters()
        except Exception as e:
            self.conn_status_label.setText("Status: Error")
            self.conn_status_label.setProperty("status", "err")
            self.conn_status_label.style().unpolish(self.conn_status_label)
            self.conn_status_label.style().polish(self.conn_status_label)

            if self.header_status is not None:
                self.header_status.setText("Disconnected")
                self.header_status.setProperty("status", "err")
                self.header_status.style().unpolish(self.header_status)
                self.header_status.style().polish(self.header_status)

            self.show_error_message(f"Device connection failed: {e}")
            self.camera = None
            self.lightbox = None

    def load_camera_parameters(self):
        if not self.camera:
            return
        try:
            # 1) PixelFormat 목록/현재값 로드
            supported_formats = self.camera.ctrl.get_enumeration_entries("PixelFormat")
            current_format = self.camera.get("PixelFormat")
            self.pixelformat_combo.clear()
            self.pixelformat_combo.addItems(supported_formats)
            if current_format in supported_formats:
                self.pixelformat_combo.setCurrentText(current_format)

            # 2) Area / Linescan 구분
            scan = ""
            with contextlib.suppress(Exception):
                scan = self.camera.get("DeviceScanType")
            if not scan:
                scan = "Linescan"
            self.is_line_camera = (str(scan) == "Linescan")

            # 3) 노출 / 라인레이트 UI 갱신
            if self.is_line_camera:
                # 라인스캔 모드 → Line Rate 사용
                self.exposure_label.setText("Line Rate (Hz):")
                lr = 0.0
                with contextlib.suppress(Exception):
                    lr = float(self.camera.get("AcquisitionLineRate"))
                if lr <= 0:
                    lr = 100000.0
                self.exposure_edit.setText(f"{lr:.1f}")

                # ✅ Dark 탭 Exp Start/End 기본값을 라인레이트 Min/Max 기반으로 설정
                self._update_dark_defaults_from_line_rate()
            else:
                # 에어리어 모드 → ExposureTime 사용(us)
                self.exposure_label.setText("Exposure (us):")
                exp = 10000.0
                with contextlib.suppress(Exception):
                    exp = float(self.camera.get("ExposureTime"))
                self.exposure_edit.setText(f"{exp:.1f}")

            # 4) Black Level
            black = 0.0
            with contextlib.suppress(Exception):
                black = float(self.camera.get("BlackLevel"))
            self.blacklevel_edit.setText(f"{black:.1f}")

        except Exception as e:
            self.show_error_message(f"Failed to load camera parameters: {e}")

    def _read_line_rate_limits_from_camera(self) -> Tuple[float, float]:
        """
        Linescan/TDI 카메라의 AcquisitionLineRate 유효 범위(min,max)를 읽는다.
        DarkDataWorker._get_line_rate_limits 와 동일 로직을 사용.
        """
        if not self.camera:
            return (30_000.0, 435_000.0)  # 최악의 경우 하드코딩 fallback

        HARD_MIN_LR = 30_000.0
        lr_min = None
        lr_max = None

        # 1) 메타데이터
        with suppress(Exception):
            meta = self.camera.ctrl.get_parameter_metadata("AcquisitionLineRate")
            if meta:
                mmin = meta.get("min")
                mmax = meta.get("max")
                if isinstance(mmin, (int, float)) and 1.0 <= mmin <= 1e9:
                    lr_min = float(mmin)
                if isinstance(mmax, (int, float)) and 1.0 <= mmax <= 1e9:
                    lr_max = float(mmax)

        # 2) Reg 노드 보조
        if lr_min is None:
            with suppress(Exception):
                vmin_reg = self.camera.get("AcquisitionLineRateMinReg")
                if vmin_reg is not None and 1.0 <= float(vmin_reg) <= 1e9:
                    lr_min = float(vmin_reg)
        if lr_max is None:
            with suppress(Exception):
                vmax_reg = self.camera.get("AcquisitionLineRateMaxReg")
                if vmax_reg is not None and 1.0 <= float(vmax_reg) <= 1e9:
                    lr_max = float(vmax_reg)

        # 3) 현재 값 기준 추정
        cur_lr = None
        with suppress(Exception):
            cur_lr = float(self.camera.get("AcquisitionLineRate"))
        if not cur_lr or cur_lr <= 0:
            cur_lr = 100_000.0

        if lr_min is None:
            lr_min = max(HARD_MIN_LR, cur_lr * 0.5)
        if lr_max is None:
            lr_max = cur_lr * 1.5

        if lr_min < HARD_MIN_LR:
            lr_min = HARD_MIN_LR
        if lr_max <= lr_min:
            lr_max = lr_min

        print(f"[DarkInit] LRmin={lr_min:.3f} Hz, LRmax={lr_max:.3f} Hz")
        return lr_min, lr_max

    def _update_dark_defaults_from_line_rate(self):
        if not self.camera or not self.is_line_camera:
            return

        # ⬇ 여기 한 줄로 min/max 얻고
        lr_min, lr_max = self._read_line_rate_limits_from_camera()

        # ⬇ TDI Stage 처리 (지금처럼 쓰고 싶으면 유지, 무시하고 1/LR 쓰고 싶으면 여기만 바꾸면 됨)
        stages: Optional[int] = None
        with suppress(Exception):
            s = self.camera.get("TDIStages")
            if isinstance(s, str):
                digits = "".join(ch for ch in s if ch.isdigit())
                if digits:
                    stages = int(digits)
            elif isinstance(s, (int, float)):
                code_map = {
                    1: 4, 2: 8, 3: 16, 4: 32, 5: 64, 6: 96, 7: 128,
                    8: 160, 9: 192, 10: 224, 11: 240, 12: 248,
                    13: 252, 14: 256,
                }
                stages = code_map.get(int(s))

        if stages is None:
            with suppress(Exception):
                h = int(self.camera.get("Height"))
                if h > 0:
                    stages = h
        if stages is None or stages <= 0:
            stages = 1

        # ==== 여기서 선택 포인트 ====
        # (A) 물리적 T_int 기준 → stages / LR
        t_int_max_raw = stages / lr_min
        t_int_min_raw = stages / lr_max

        # (B) TDI stage 무시하고 1/LR만 쓰고 싶으면 위 둘 대신:
        # t_int_max_raw = 1.0 / lr_min
        # t_int_min_raw = 1.0 / lr_max
        # ===========================

        if t_int_max_raw <= 0 or t_int_min_raw <= 0:
            return
        if t_int_max_raw < t_int_min_raw:
            t_int_min_raw, t_int_max_raw = t_int_max_raw, t_int_min_raw

        t_int_max = t_int_max_raw
        t_int_min = max(t_int_min_raw, t_int_max / 100.0, 1e-6)

        print(
            f"[DarkInit] LRmin={lr_min:.3f} Hz, LRmax={lr_max:.3f} Hz, "
            f"stages={stages}, T_int_min={t_int_min:.6e} s, T_int_max={t_int_max:.6e} s"
        )

        self.dark_exp_start_edit.setText(f"{t_int_min:.6f}")
        self.dark_exp_end_edit.setText(f"{t_int_max:.6f}")


    def apply_camera_parameters(self):
        if not self.camera:
            return self.show_error_message("Camera not connected.")
        was_live = self.live_view_btn.isChecked()
        if was_live:
            self.live_timer.stop()
            with contextlib.suppress(Exception):
                self.camera.stop_preview()
        try:
            pf = self.pixelformat_combo.currentText()
            value = float(self.exposure_edit.text())
            black = int(float(self.blacklevel_edit.text()))
            self.camera.set("PixelFormat", pf)
            if self.is_line_camera:
                self.camera.set("AcquisitionLineRate", value)
            else:
                self.camera.set("ExposureTime", value)
            self.camera.set("BlackLevel", black)
            if hasattr(self.camera, "flush_stream"):
                self.camera.flush_stream()
            time.sleep(0.02)
            curp = self.bayer_pattern_combo.currentText()
            mode = f"Applied. PF={pf}, Pattern={curp} → {'G-only(compact+upsample)' if curp in VALID_BAYER else 'Original'}"
            self.status_label_bottom.setText(mode)
            self.statusBar().showMessage(mode)
        except Exception as e:
            self.show_error_message(f"Failed to apply parameters: {e}")
        finally:
            if was_live:
                with contextlib.suppress(Exception):
                    self.camera.start_preview()
                self.live_timer.start(33)
                self.update_live_view()

    @pyqtSlot(bool)
    def toggle_live_view(self, checked: bool):
        if not self.camera:
            btns = [self.live_view_btn]
            if self.quick_live_btn is not None:
                btns.append(self.quick_live_btn)
            for b in btns:
                with contextlib.suppress(Exception):
                    b.setChecked(False)
            self.show_error_message("Camera not connected.")
            return

        def _sync(on: bool):
            btns = [self.live_view_btn]
            if self.quick_live_btn is not None:
                btns.append(self.quick_live_btn)
            for b in btns:
                if b.isChecked() != on:
                    b.blockSignals(True)
                    b.setChecked(on)
                    b.blockSignals(False)
                b.setText("Stop Live" if on else "Start Live")
                b.setIcon(self.style().standardIcon(QStyle.SP_MediaStop if on else QStyle.SP_MediaPlay))

        if checked:
            try:
                self.apply_camera_parameters()
                if hasattr(self.camera, "start_preview"):
                    self.camera.start_preview()
                self.live_timer.start(33)
                _sync(True)
                self.statusBar().showMessage("Live view started.")
            except Exception as e:
                self.live_timer.stop()
                with contextlib.suppress(Exception):
                    if hasattr(self.camera, "stop_preview"):
                        self.camera.stop_preview()
                _sync(False)
                self.show_error_message(f"Failed to start live view: {e}")
        else:
            try:
                self.live_timer.stop()
                with contextlib.suppress(Exception):
                    if hasattr(self.camera, "stop_preview"):
                        self.camera.stop_preview()
                time.sleep(0.05)
                _sync(False)
                self.status_label_bottom.setText("Live view stopped.")
                self.statusBar().showMessage("Live view stopped.")
            except Exception as e:
                _sync(False)
                self.live_timer.stop()
                self.show_error_message(f"Failed to stop live view: {e}")


# -------------------------------
# Entry
# -------------------------------
if __name__ == '__main__':
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)

    # ★ AUMID: Qt 생성 전에!
    _set_windows_app_user_model_id(AUMID)

    args = parse_args()

    app = QApplication(sys.argv)
    app.setStyle(CompactProxyStyle("Fusion"))

    # Theme
    if args.theme == "auto":
        _apply_dark(app) if detect_windows_dark() else _apply_light(app)
    elif args.theme == "dark":
        _apply_dark(app)
    else:
        _apply_light(app)

    # Icon (한 번만 로드) — 기본 vieworks.ico 우선
    app_icon, icon_path = load_app_icon(args.icon)
    if app_icon:
        app.setWindowIcon(app_icon)

    ex = EMVAAnalyzerApp(
        forced_icon_path=icon_path if icon_path else None,
        show_header=False,  # ← 헤더 항상 숨김
        # frameless=args.frameless  # 프레임리스 쓰는 중이면 유지
    )

    if app_icon:
        ex.setWindowIcon(app_icon)  # 일부 환경 안정성 확보

    ex.show()
    sys.exit(app.exec_())

